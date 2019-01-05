# -*- coding: utf-8 -*-
"""
Created on Sat Aug 18 16:18:35 2018

@author: shiva
"""

#import PyPDF2 
from nltk.tokenize import word_tokenize
from nltk.corpus import stopwords
import datetime,calendar
from datetime import timedelta
import xlrd
import openpyxl
from pyjarowinkler import distance

from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage
from io import StringIO

from termcolor import colored

book = openpyxl.load_workbook('CitiesDataExcelSheet.xlsx')

cityNamesFromSheet = book.get_sheet_names()

index = 0


def convert_pdf_to_txt(path):
    rsrcmgr = PDFResourceManager()
    retstr = StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    fp = open(path, 'rb')
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    maxpages = 0
    caching = True
    pagenos=set()

    for page in PDFPage.get_pages(fp, pagenos, maxpages=maxpages, password=password,caching=caching, check_extractable=True):
        interpreter.process_page(page)

    text = retstr.getvalue()

    fp.close()
    device.close()
    retstr.close()
    return text


def writingPDFtoXL(cityName):
    
    global index
    #Applying Jaro winkler similarity frm here    
    XLactiveSheetName = cityNamesFromSheet[index]   
    index = index + 1    
    similarity = distance.get_jaro_distance(XLactiveSheetName, cityName, winkler=True, scaling=0.1)    
    if similarity >= 0.0:
        cityName = XLactiveSheetName
        print("Accepted++" ,similarity)
        print("XLactiveSheetName--",XLactiveSheetName)
        print("cityName--",cityName)
        print ("*_*__*_*_*_*_*_*_*")
    else:
        print("Rejected--",similarity)
        print("XLactiveSheetName--",XLactiveSheetName)
        print("cityName--",cityName)
        print ("*_*__*_*_*_*_*_*_*")
        return
    
    
#==============================================================================
#     pdfFileObj = open('somepdf.pdf','rb')
#     #The pdfReader variable is a readable object that will be parsed
#     pdfReader = PyPDF2.PdfFileReader(pdfFileObj)
#     
#     num_pages = pdfReader.numPages
#     count = 0
#     text = ""
#     
#     #The while loop will read each page
#     while count < num_pages:
#         pageObj = pdfReader.getPage(count)
#         count +=1
#         text += pageObj.extractText()
#==============================================================================
    text = convert_pdf_to_txt('somepdf.pdf')

    #This if statement exists to check if the above library returned #words. It's done because PyPDF2 cannot read scanned files.
    if text != "":
       text = text
    #If the above returns as False, we run the OCR library textract to #convert scanned/image based PDF files into text
    else:
        print ('It is scanned PDF.')
    #   text = textract.process(fileurl, method='tesseract', language='eng')
    
    #The word_tokenize() function will break our text phrases into #individual words
    tokens = word_tokenize(text)
    #we'll create a new list which contains punctuation we wish to clean
    punctuations = ['(',')',';',':','[',']',',','.']
    #We initialize the stopwords variable which is a list of words like #"The", "I", "and", etc. that don't hold much value as keywords
    stop_words = stopwords.words('english')
    #We create a list comprehension which only returns a list of words #that are NOT IN stop_words and NOT IN punctuations.
    keywords = [word for word in tokens if not word in stop_words and not word in punctuations]
    startNumber = keywords[keywords.index('Starting')+2] #Here Starting is the keyword from PDF see PDF for Starting word
    

    
    #Create work book
    bookXLRD = xlrd.open_workbook('CitiesDataExcelSheet.xlsx')
    sheetXLRD = bookXLRD.sheet_by_name(cityName)
    
    book = openpyxl.load_workbook('CitiesDataExcelSheet.xlsx')
    
    
    
    
    activeSheetName = book.get_sheet_by_name(cityName)

    
    try:
        rowNo = sheetXLRD.nrows + 1
        for word in keywords:
            if word == 'Starting':        
                activeSheetName.cell(row = rowNo, column = 1).value = datetime.date.today()
                startNumber = keywords[keywords.index('Starting')+1]
                endNumber = keywords[keywords.index('Ending')+2]
                activeSheetName.cell(row = rowNo, column = 2).value = startNumber[3:]
                activeSheetName.cell(row = rowNo, column = 3).value = endNumber
                activeSheetName.cell(row = rowNo, column = 4).value = int(startNumber[-4:])
                activeSheetName.cell(row = rowNo, column = 5).value = calendar.day_name[datetime.date.today().weekday()]
                
                if datetime.datetime.strftime(activeSheetName.cell(row = rowNo - 1, column = 1).value, '%Y-%m-%d') ==  datetime.datetime.strftime(activeSheetName.cell(row = rowNo, column = 1).value - timedelta(1), '%Y-%m-%d'):
                    noOfVehicles = activeSheetName.cell(row = rowNo, column = 4).value - activeSheetName.cell(row = rowNo -1, column = 4).value
                    
                    if noOfVehicles >= 0:
                        activeSheetName.cell(row = rowNo, column = 6).value = noOfVehicles
                    else:
                        activeSheetName.cell(row = rowNo, column = 6).value = (9999 - abs(noOfVehicles) + activeSheetName.cell(row = rowNo, column = 4).value)
                else:
                    print('Failed IF')
                

                
                rowNo += 1
    except:
        print(colored('some error in pdf parsing' ,'red' ))
        pass
    book.save('CitiesDataExcelSheet.xlsx')
    print ('Saved'+ '--' + cityName)
    
    print (XLactiveSheetName)
    print (cityName)
    print('-------------------------------------------------')

